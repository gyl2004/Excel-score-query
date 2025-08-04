"""
报告生成器测试
"""
import unittest
import tempfile
import os
from pathlib import Path
import openpyxl

from services.report_generator import ReportGenerator
from models.data_models import PositionScoreResult


class TestReportGenerator(unittest.TestCase):
    """报告生成器测试类"""
    
    def setUp(self):
        """测试前准备"""
        self.generator = ReportGenerator()
        self.temp_dir = tempfile.mkdtemp()
        
        # 创建测试数据
        self.test_results = [
            PositionScoreResult(
                position_code="P001",
                position_name="软件工程师",
                department="技术部",
                min_score=85.5,
                candidate_count=5,
                status="正常",
                notes=""
            ),
            PositionScoreResult(
                position_code="P002",
                position_name="产品经理",
                department="产品部",
                min_score=None,
                candidate_count=0,
                status="无面试人员",
                notes="该岗位暂无面试人员"
            ),
            PositionScoreResult(
                position_code="P003",
                position_name="UI设计师",
                department="设计部",
                min_score=78.0,
                candidate_count=3,
                status="正常",
                notes=""
            )
        ]
    
    def tearDown(self):
        """测试后清理"""
        # 清理临时文件
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def test_generate_report_success(self):
        """测试成功生成报告"""
        output_path = os.path.join(self.temp_dir, "test_report.xlsx")
        
        success, final_path = self.generator.generate_report(self.test_results, output_path)
        
        self.assertTrue(success)
        self.assertTrue(os.path.exists(final_path))
        self.assertEqual(final_path, os.path.abspath(output_path))
        
        # 验证文件可以正常打开
        wb = openpyxl.load_workbook(final_path)
        ws = wb.active
        
        # 验证标题
        self.assertEqual(ws['A1'].value, "岗位最低进面分数汇总报告")
        
        # 验证表头
        expected_headers = ['岗位代码', '岗位名称', '部门', '最低分数', '面试人数', '状态', '备注']
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=4, column=col).value
            self.assertEqual(actual_header, expected_header)
        
        # 验证数据行
        self.assertEqual(ws.cell(row=5, column=1).value, "P001")
        self.assertEqual(ws.cell(row=5, column=2).value, "软件工程师")
        self.assertEqual(ws.cell(row=5, column=4).value, 85.5)
        
        # 验证无数据情况
        self.assertEqual(ws.cell(row=6, column=4).value, "无数据")
        self.assertEqual(ws.cell(row=6, column=6).value, "无面试人员")
        
        wb.close()
    
    def test_generate_report_empty_results(self):
        """测试空结果列表"""
        output_path = os.path.join(self.temp_dir, "empty_report.xlsx")
        
        success, final_path = self.generator.generate_report([], output_path)
        
        self.assertTrue(success)
        self.assertTrue(os.path.exists(final_path))
        
        # 验证文件结构
        wb = openpyxl.load_workbook(final_path)
        ws = wb.active
        
        # 应该有标题和表头，但没有数据行
        self.assertEqual(ws['A1'].value, "岗位最低进面分数汇总报告")
        self.assertIsNotNone(ws.cell(row=4, column=1).value)  # 表头应该存在
        
        wb.close()
    
    def test_generate_report_invalid_path(self):
        """测试无效输出路径"""
        invalid_path = "Z:\\nonexistent\\path\\report.xlsx"  # 使用不存在的驱动器
        
        with self.assertRaises(Exception) as context:
            self.generator.generate_report(self.test_results, invalid_path)
        
        self.assertIn("生成报告失败", str(context.exception))
    
    def test_create_summary_statistics(self):
        """测试汇总统计信息"""
        stats = self.generator.create_summary_statistics(self.test_results)
        
        self.assertEqual(stats["total_positions"], 3)
        self.assertEqual(stats["normal_positions"], 2)
        self.assertEqual(stats["no_interview_positions"], 1)
        self.assertEqual(stats["error_positions"], 0)
        self.assertEqual(stats["success_rate"], "66.7%")
    
    def test_create_summary_statistics_empty(self):
        """测试空结果的汇总统计"""
        stats = self.generator.create_summary_statistics([])
        
        self.assertEqual(stats["total_positions"], 0)
        self.assertEqual(stats["success_rate"], "0%")
    
    def test_report_header_contains_timestamp(self):
        """测试报告包含时间戳"""
        output_path = os.path.join(self.temp_dir, "timestamp_report.xlsx")
        
        success, final_path = self.generator.generate_report(self.test_results, output_path)
        
        wb = openpyxl.load_workbook(final_path)
        ws = wb.active
        
        # 验证时间戳格式
        time_cell_value = ws['A2'].value
        self.assertIsNotNone(time_cell_value)
        self.assertIn("生成时间:", time_cell_value)
        
        wb.close()
    
    def test_report_styling_applied(self):
        """测试报告样式应用"""
        output_path = os.path.join(self.temp_dir, "styled_report.xlsx")
        
        success, final_path = self.generator.generate_report(self.test_results, output_path)
        
        wb = openpyxl.load_workbook(final_path)
        ws = wb.active
        
        # 验证标题样式
        title_cell = ws['A1']
        self.assertTrue(title_cell.font.bold)
        self.assertEqual(title_cell.font.size, 16)
        
        # 验证表头样式
        header_cell = ws.cell(row=4, column=1)
        self.assertTrue(header_cell.font.bold)
        self.assertEqual(header_cell.font.color.rgb, '00FFFFFF')  # 白色字体
        
        wb.close()
    
    def test_generate_report_file_conflict_auto_rename(self):
        """测试文件冲突自动重命名"""
        output_path = os.path.join(self.temp_dir, "conflict_report.xlsx")
        
        # 先创建一个同名文件
        with open(output_path, 'w') as f:
            f.write("existing file")
        
        success, final_path = self.generator.generate_report(self.test_results, output_path, "auto_rename")
        
        self.assertTrue(success)
        self.assertNotEqual(final_path, output_path)
        self.assertTrue(final_path.endswith("_1.xlsx"))
        self.assertTrue(os.path.exists(final_path))
    
    def test_generate_report_file_conflict_overwrite(self):
        """测试文件冲突覆盖"""
        output_path = os.path.join(self.temp_dir, "overwrite_report.xlsx")
        
        # 先创建一个同名文件
        with open(output_path, 'w') as f:
            f.write("existing file")
        
        success, final_path = self.generator.generate_report(self.test_results, output_path, "overwrite")
        
        self.assertTrue(success)
        self.assertEqual(final_path, os.path.abspath(output_path))
        self.assertTrue(os.path.exists(final_path))
    
    def test_get_recommended_save_path(self):
        """测试获取推荐保存路径"""
        path = self.generator.get_recommended_save_path()
        
        self.assertTrue(path.endswith("岗位最低分数汇总.xlsx"))
        
        # 测试指定目录
        custom_path = self.generator.get_recommended_save_path(self.temp_dir)
        expected_path = os.path.join(self.temp_dir, "岗位最低分数汇总.xlsx")
        self.assertEqual(custom_path, os.path.abspath(expected_path))
    
    def test_validate_output_path(self):
        """测试验证输出路径"""
        valid_path = os.path.join(self.temp_dir, "valid.xlsx")
        is_valid, error_msg = self.generator.validate_output_path(valid_path)
        
        self.assertTrue(is_valid)
        self.assertEqual(error_msg, "")
        
        # 测试无效路径
        invalid_path = "Z:\\nonexistent\\path\\file.xlsx"  # 使用不存在的驱动器
        is_valid, error_msg = self.generator.validate_output_path(invalid_path)
        
        self.assertFalse(is_valid)
        self.assertNotEqual(error_msg, "")


if __name__ == '__main__':
    unittest.main()