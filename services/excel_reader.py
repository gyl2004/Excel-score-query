"""
Excel文件读取器
提供Excel文件读取和验证功能
"""
import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional
import logging


class ExcelProcessingError(Exception):
    """Excel文件处理相关异常"""
    pass


class ExcelReader:
    """Excel文件读取器类"""
    
    def __init__(self):
        """初始化Excel读取器"""
        self.logger = logging.getLogger(__name__)
        self.supported_extensions = ['.xlsx', '.xls']
    
    def get_column_names(self, file_path: str, sheet_name: Optional[str] = None) -> List[str]:
        """
        获取Excel文件的列名，智能检测标题行位置
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，如果为None则使用第一个工作表
            
        Returns:
            列名列表
            
        Raises:
            ExcelProcessingError: 读取失败时抛出异常
        """
        try:
            self.logger.info(f"读取Excel文件列名: {file_path}")
            
            # 验证文件
            self.validate_file_path(file_path)
            
            # 智能检测标题行
            columns = self._detect_header_row(file_path, sheet_name)
            
            if not columns:
                raise ExcelProcessingError("未能读取到有效的列名")
            
            self.logger.info(f"成功读取列名: {columns}")
            return columns
            
        except Exception as e:
            if isinstance(e, ExcelProcessingError):
                raise
            else:
                error_msg = f"读取Excel文件列名失败: {str(e)}"
                self.logger.error(error_msg)
                raise ExcelProcessingError(error_msg)
    
    def _detect_header_row_index(self, file_path: str, sheet_name: Optional[str] = None) -> int:
        """
        智能检测Excel文件的标题行索引位置
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            标题行的索引位置（0-based）
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # 读取前5行数据来检测标题行
            rows_data = []
            max_cols = 0
            
            # 先确定最大列数
            for row in ws.iter_rows(min_row=1, max_row=5):
                max_cols = max(max_cols, len(row))
            
            for row_num in range(1, 6):  # 读取前5行
                row_data = []
                
                # 确保读取所有列
                for col_num in range(1, max_cols + 1):
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        cell_value = cell.value
                        if cell_value is not None:
                            row_data.append(str(cell_value).strip())
                        else:
                            row_data.append('')
                    except:
                        row_data.append('')
                
                rows_data.append(row_data)
            
            wb.close()
            
            # 分析每一行，找到最可能是标题行的行
            best_header_row_index = 0
            best_score = -1
            
            for row_idx, row_data in enumerate(rows_data):
                if not row_data or all(not cell for cell in row_data):
                    continue  # 跳过空行
                
                score = self._calculate_header_score(row_data)
                self.logger.debug(f"第{row_idx+1}行标题分数: {score}, 内容: {row_data[:3]}...")
                
                if score > best_score:
                    best_score = score
                    best_header_row_index = row_idx
            
            return best_header_row_index
            
        except Exception as e:
            self.logger.warning(f"智能检测标题行位置失败: {e}，使用默认位置0")
            return 0

    def _detect_header_row(self, file_path: str, sheet_name: Optional[str] = None) -> List[str]:
        """
        智能检测Excel文件的标题行位置
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            列名列表
        """
        # 首先尝试使用pandas直接读取（处理合并单元格等复杂情况）
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
            else:
                df = pd.read_excel(file_path, nrows=0)
            
            columns = [str(col).strip() for col in df.columns 
                      if str(col).strip() and str(col) != 'nan' 
                      and not str(col).startswith('Unnamed:')]
            
            if len(columns) >= 3:  # 如果pandas能读取到足够的列名，直接使用
                self.logger.debug(f"pandas成功读取列名: {columns}")
                return columns
                
        except Exception as e:
            self.logger.debug(f"pandas读取失败: {e}")
        
        # 如果pandas失败，使用openpyxl进行智能检测
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # 读取前5行数据来检测标题行
            rows_data = []
            max_cols = 0
            
            # 先确定最大列数
            for row in ws.iter_rows(min_row=1, max_row=5):
                max_cols = max(max_cols, len(row))
            
            for row_num in range(1, 6):  # 读取前5行
                row_data = []
                
                # 确保读取所有列
                for col_num in range(1, max_cols + 1):
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        cell_value = cell.value
                        if cell_value is not None:
                            row_data.append(str(cell_value).strip())
                        else:
                            row_data.append('')
                    except:
                        row_data.append('')
                
                rows_data.append(row_data)
            
            wb.close()
            
            # 分析每一行，找到最可能是标题行的行
            best_header_row = None
            best_score = -1
            
            for row_idx, row_data in enumerate(rows_data):
                if not row_data or all(not cell for cell in row_data):
                    continue  # 跳过空行
                
                score = self._calculate_header_score(row_data)
                self.logger.debug(f"第{row_idx+1}行标题分数: {score}, 内容: {row_data[:3]}...")
                
                if score > best_score:
                    best_score = score
                    best_header_row = row_data
            
            if best_header_row:
                # 清理列名
                cleaned_columns = []
                for col in best_header_row:
                    col_str = str(col).strip()
                    if col_str and col_str != 'nan' and col_str != 'None' and not col_str.startswith('Unnamed:'):
                        cleaned_columns.append(col_str)
                
                if cleaned_columns:
                    return cleaned_columns
            
            # 如果智能检测失败，回退到传统方法
            self.logger.warning("智能检测失败，使用传统方法读取列名")
            return self._fallback_column_reading(file_path, sheet_name)
            
        except Exception as e:
            self.logger.warning(f"智能检测失败: {e}，使用传统方法")
            return self._fallback_column_reading(file_path, sheet_name)
    
    def _calculate_header_score(self, row_data: List[str]) -> float:
        """
        计算一行数据作为标题行的可能性分数
        
        Args:
            row_data: 行数据
            
        Returns:
            分数，越高越可能是标题行
        """
        if not row_data:
            return 0
        
        score = 0
        non_empty_count = 0
        
        # 常见的标题关键词（增加更多关键词）
        header_keywords = [
            '准考证', '姓名', '部门', '职位', '岗位', '代码', '分数', '成绩',
            '招录', '招考', '用人', '司局', '机关', '单位', '最低', '面试',
            '部门代码', '部门名称', '用人司局', '机构性质', '招考职位', 
            '职位属性', '职位分布', '职位简介', '职位代码', '机构层级',
            '考试类别', '招考人数', '专业', '学历', '学位', '政治面貌'
        ]
        
        # 检查是否包含必需的关键列（岗位表）
        required_position_keywords = ['用人司局', '招考职位', '职位代码']
        required_interview_keywords = ['姓名', '岗位', '分数']
        
        for cell in row_data:
            if not cell:
                continue
            
            non_empty_count += 1
            cell_str = str(cell).strip()
            
            # 检查是否包含必需的岗位表关键词（高分奖励）
            for keyword in required_position_keywords:
                if keyword in cell_str:
                    score += 50  # 大幅提高分数
                    break
            
            # 检查是否包含必需的面试表关键词（高分奖励）
            for keyword in required_interview_keywords:
                if keyword in cell_str:
                    score += 30
                    break
            
            # 检查是否包含其他标题关键词
            for keyword in header_keywords:
                if keyword in cell_str:
                    score += 10
                    break
            
            # 检查长度（标题通常比较短）
            if 2 <= len(cell_str) <= 15:
                score += 3
            elif len(cell_str) > 50:  # 很长的文本通常不是标题
                score -= 20
            
            # 检查是否包含数字（标题通常不全是数字）
            if cell_str.isdigit():
                score -= 8
            
            # 检查是否是说明性文字（大幅降分）
            if any(word in cell_str for word in ['说明', '注意', '备注', '由', '负责', '解释', '编报', '相关的问题']):
                score -= 30
            
            # 检查是否包含"Unnamed"（pandas默认列名，降分）
            if 'Unnamed' in cell_str:
                score -= 20
        
        # 非空单元格数量奖励
        if non_empty_count >= 5:  # 岗位表通常有很多列
            score += non_empty_count * 3
        elif non_empty_count >= 3:  # 面试表通常有较少列
            score += non_empty_count * 2
        
        # 如果包含多个必需关键词，额外奖励
        position_keyword_count = sum(1 for keyword in required_position_keywords 
                                   if any(keyword in str(cell) for cell in row_data if cell))
        if position_keyword_count >= 2:
            score += position_keyword_count * 20
        
        return score
    
    def _fallback_column_reading(self, file_path: str, sheet_name: Optional[str] = None) -> List[str]:
        """
        传统的列名读取方法（回退方案）
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            列名列表
        """
        try:
            # 尝试使用pandas读取，跳过可能的说明行
            for skip_rows in [0, 1, 2]:  # 尝试跳过0、1、2行
                try:
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, 
                                         skiprows=skip_rows, nrows=1)
                    else:
                        df = pd.read_excel(file_path, skiprows=skip_rows, nrows=1)
                    
                    columns = [str(col).strip() for col in df.columns 
                             if str(col).strip() and str(col) != 'nan' 
                             and not str(col).startswith('Unnamed:')]
                    
                    if len(columns) >= 3:  # 至少要有3个有效列名
                        return columns
                        
                except Exception:
                    continue
            
            # 如果还是失败，返回空列表
            return []
            
        except Exception as e:
            self.logger.error(f"传统方法读取列名失败: {e}")
            return []
        
    def validate_file_path(self, file_path: str) -> bool:
        """
        验证文件路径是否有效
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 文件路径是否有效
            
        Raises:
            ExcelProcessingError: 文件路径无效时抛出异常
        """
        if not file_path:
            raise ExcelProcessingError("文件路径不能为空")
            
        path = Path(file_path)
        
        if not path.exists():
            raise ExcelProcessingError(f"文件不存在: {file_path}")
            
        if not path.is_file():
            raise ExcelProcessingError(f"路径不是文件: {file_path}")
            
        # 检查文件扩展名
        valid_extensions = {'.xlsx', '.xls', '.xlsm'}
        if path.suffix.lower() not in valid_extensions:
            raise ExcelProcessingError(f"不支持的文件格式: {path.suffix}，支持的格式: {', '.join(valid_extensions)}")
            
        # 检查文件是否可读
        if not os.access(file_path, os.R_OK):
            raise ExcelProcessingError(f"文件无法读取，请检查文件权限: {file_path}")
            
        return True
        
    def check_excel_format(self, file_path: str) -> bool:
        """
        检查Excel文件格式是否正确
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            bool: 文件格式是否正确
            
        Raises:
            ExcelProcessingError: 文件格式错误时抛出异常
        """
        try:
            # 尝试读取Excel文件的基本信息
            excel_file = pd.ExcelFile(file_path)
            
            # 检查是否有sheet
            if not excel_file.sheet_names:
                raise ExcelProcessingError(f"Excel文件没有工作表: {file_path}")
                
            # 检查每个sheet是否可以读取
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1)
                    if df.empty:
                        self.logger.warning(f"工作表 '{sheet_name}' 为空")
                except Exception as e:
                    raise ExcelProcessingError(f"无法读取工作表 '{sheet_name}': {str(e)}")
                    
            excel_file.close()
            return True
            
        except pd.errors.EmptyDataError:
            raise ExcelProcessingError(f"Excel文件为空: {file_path}")
        except pd.errors.ParserError as e:
            raise ExcelProcessingError(f"Excel文件格式错误: {str(e)}")
        except Exception as e:
            raise ExcelProcessingError(f"读取Excel文件时发生错误: {str(e)}")
            
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        获取Excel文件中所有工作表名称
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            List[str]: 工作表名称列表
            
        Raises:
            ExcelProcessingError: 读取失败时抛出异常
        """
        self.validate_file_path(file_path)
        
        try:
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            excel_file.close()
            
            self.logger.info(f"成功获取工作表名称，共 {len(sheet_names)} 个: {sheet_names}")
            return sheet_names
            
        except Exception as e:
            raise ExcelProcessingError(f"获取工作表名称失败: {str(e)}")
            
    def read_excel_sheet(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        读取Excel文件的指定工作表
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，如果为None则读取第一个工作表
            
        Returns:
            pd.DataFrame: 读取的数据
            
        Raises:
            ExcelProcessingError: 读取失败时抛出异常
        """
        self.validate_file_path(file_path)
        self.check_excel_format(file_path)
        
        try:
            # 如果没有指定sheet名称，使用第一个sheet
            if sheet_name is None:
                sheet_names = self.get_sheet_names(file_path)
                sheet_name = sheet_names[0] if sheet_names else None
                
            if sheet_name is None:
                raise ExcelProcessingError("无法确定要读取的工作表")
                
            # 使用智能检测来确定正确的标题行位置
            header_row_index = self._detect_header_row_index(file_path, sheet_name)
            
            # 定义特殊列的数据类型，确保职位代码等关键字段以文本格式读取
            dtype_dict = {}
            
            # 预读取列名以确定数据类型
            try:
                if header_row_index > 0:
                    temp_df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=header_row_index, nrows=0)
                else:
                    temp_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
                
                # 为包含"代码"的列设置文本类型
                for col in temp_df.columns:
                    col_str = str(col).strip().lower()
                    if any(keyword in col_str for keyword in ['代码', 'code', '编号', '号码']):
                        dtype_dict[col] = str
                        self.logger.debug(f"设置列 '{col}' 为文本类型")
                        
            except Exception as e:
                self.logger.warning(f"预读取列名失败，使用默认数据类型: {e}")
            
            # 读取数据，使用检测到的标题行和数据类型
            if header_row_index > 0:
                df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=header_row_index, dtype=dtype_dict)
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=dtype_dict)
            
            # 基本数据清理
            df = df.dropna(how='all')  # 删除完全为空的行
            df = df.dropna(axis=1, how='all')  # 删除完全为空的列
            
            # 后处理：确保关键列为文本格式
            for col in df.columns:
                col_str = str(col).strip().lower()
                if any(keyword in col_str for keyword in ['代码', 'code', '编号', '号码']):
                    df[col] = df[col].astype(str)
                    # 清理可能的科学计数法表示
                    df[col] = df[col].apply(self._clean_scientific_notation)
            
            self.logger.info(f"成功读取工作表 '{sheet_name}'，数据行数: {len(df)}")
            return df
            
        except FileNotFoundError:
            raise ExcelProcessingError(f"文件不存在: {file_path}")
        except ValueError as e:
            if "Worksheet" in str(e) and "does not exist" in str(e):
                raise ExcelProcessingError(f"工作表 '{sheet_name}' 不存在")
            else:
                raise ExcelProcessingError(f"读取Excel文件时发生错误: {str(e)}")
        except Exception as e:
            raise ExcelProcessingError(f"读取Excel文件时发生未知错误: {str(e)}")
    
    def _clean_scientific_notation(self, value):
        """
        清理科学计数法表示，转换为正常数字字符串
        
        Args:
            value: 输入值
            
        Returns:
            清理后的字符串
        """
        try:
            if pd.isna(value) or value == '' or str(value).lower() == 'nan':
                return ''
            
            value_str = str(value).strip()
            
            # 检查是否是科学计数法格式
            if 'E' in value_str.upper() or 'e' in value_str:
                try:
                    # 尝试转换为浮点数再转为整数字符串
                    float_val = float(value_str)
                    if float_val.is_integer():
                        return str(int(float_val))
                    else:
                        return str(float_val)
                except (ValueError, OverflowError):
                    pass
            
            return value_str
            
        except Exception:
            return str(value) if value is not None else ''
            
    def read_all_sheets(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        读取Excel文件的所有工作表
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict[str, pd.DataFrame]: 工作表名称到数据框的映射
            
        Raises:
            ExcelProcessingError: 读取失败时抛出异常
        """
        self.validate_file_path(file_path)
        self.check_excel_format(file_path)
        
        try:
            sheet_names = self.get_sheet_names(file_path)
            sheets_data = {}
            
            for sheet_name in sheet_names:
                try:
                    df = self.read_excel_sheet(file_path, sheet_name)
                    sheets_data[sheet_name] = df
                    self.logger.info(f"成功读取工作表 '{sheet_name}'")
                except Exception as e:
                    self.logger.error(f"读取工作表 '{sheet_name}' 失败: {str(e)}")
                    # 继续读取其他工作表，不中断整个过程
                    
            if not sheets_data:
                raise ExcelProcessingError("没有成功读取任何工作表")
                
            return sheets_data
            
        except Exception as e:
            if isinstance(e, ExcelProcessingError):
                raise
            else:
                raise ExcelProcessingError(f"读取所有工作表时发生错误: {str(e)}")
                
    def validate_position_data_structure(self, df: pd.DataFrame, sheet_name: str) -> bool:
        """
        验证职位表数据结构（在预处理之前调用）
        
        Args:
            df: 原始数据框
            sheet_name: 工作表名称
            
        Returns:
            bool: 数据结构是否有效
            
        Raises:
            ExcelProcessingError: 数据结构无效时抛出异常
        """
        if df.empty:
            raise ExcelProcessingError(f"工作表 '{sheet_name}' 没有数据")
            
        # 检查必需的列（灵活匹配列名）
        required_keywords = ['岗位', '职位']
        position_column = None
        
        for col in df.columns:
            col_str = str(col).strip()
            if any(keyword in col_str for keyword in required_keywords):
                position_column = col
                break
                
        if position_column is None:
            available_columns = list(df.columns)
            raise ExcelProcessingError(
                f"工作表 '{sheet_name}' 缺少职位相关列。"
                f"需要包含包含以下关键词的列: {required_keywords}，"
                f"当前列: {available_columns}"
            )
            
        # 检查职位列是否有有效数据
        valid_positions = df[position_column].dropna()
        if len(valid_positions) == 0:
            raise ExcelProcessingError(f"工作表 '{sheet_name}' 的职位列没有有效数据")
            
        self.logger.info(f"工作表 '{sheet_name}' 数据结构验证通过，找到 {len(valid_positions)} 个有效职位")
        return True
        
    def read_position_file(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        读取职位表文件，支持多sheet处理
        
        Args:
            file_path: 职位表文件路径
            
        Returns:
            Dict[str, pd.DataFrame]: 工作表名称到职位数据的映射
            
        Raises:
            ExcelProcessingError: 读取失败时抛出异常
        """
        self.validate_file_path(file_path)
        self.check_excel_format(file_path)
        
        try:
            # 读取所有工作表
            all_sheets = self.read_all_sheets(file_path)
            position_sheets = {}
            
            for sheet_name, df in all_sheets.items():
                try:
                    # 验证每个工作表的数据结构
                    self.validate_position_data_structure(df, sheet_name)
                    
                    # 数据预处理
                    processed_df = self._preprocess_position_data(df, sheet_name)
                    position_sheets[sheet_name] = processed_df
                    
                    self.logger.info(f"成功处理职位表工作表 '{sheet_name}'，包含 {len(processed_df)} 个职位")
                    
                except ExcelProcessingError as e:
                    self.logger.warning(f"跳过工作表 '{sheet_name}': {str(e)}")
                    # 继续处理其他工作表
                    
            if not position_sheets:
                raise ExcelProcessingError("没有找到有效的职位数据工作表")
                
            self.logger.info(f"职位表读取完成，共处理 {len(position_sheets)} 个工作表")
            return position_sheets
            
        except Exception as e:
            if isinstance(e, ExcelProcessingError):
                raise
            else:
                raise ExcelProcessingError(f"读取职位表文件时发生错误: {str(e)}")
                
    def _preprocess_position_data(self, df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """
        预处理职位数据
        
        Args:
            df: 原始数据框
            sheet_name: 工作表名称
            
        Returns:
            pd.DataFrame: 预处理后的数据框
        """
        # 创建数据副本
        processed_df = df.copy()
        
        # 保留所有原始列名，同时添加标准化的别名列
        # 这样既保持了原始列名（供可配置匹配器使用），又有标准化的列名（供传统匹配器使用）
        
        # 查找并添加标准化的position_name列
        position_name_col = None
        for col in processed_df.columns:
            col_str = str(col).strip()
            if any(keyword in col_str for keyword in ['招考职位', '岗位名称', '职位名称']):
                position_name_col = col
                break
        
        if position_name_col:
            # 添加标准化的position_name列（如果不存在的话）
            if 'position_name' not in processed_df.columns:
                processed_df['position_name'] = processed_df[position_name_col]
        else:
            # 如果没有找到明确的职位名称列，使用第一个包含职位信息的列
            for col in processed_df.columns:
                col_str = str(col).strip()
                if any(keyword in col_str for keyword in ['岗位', '职位']):
                    processed_df['position_name'] = processed_df[col]
                    break
        
        # 查找并添加标准化的position_code列
        position_code_col = None
        for col in processed_df.columns:
            col_str = str(col).strip()
            if any(keyword in col_str for keyword in ['职位代码', '岗位代码']):
                position_code_col = col
                break
        
        if position_code_col and 'position_code' not in processed_df.columns:
            processed_df['position_code'] = processed_df[position_code_col]
        
        # 查找并添加标准化的department列
        department_col = None
        for col in processed_df.columns:
            col_str = str(col).strip()
            if any(keyword in col_str for keyword in ['部门名称', '用人司局']):
                department_col = col
                break
        
        if department_col and 'department' not in processed_df.columns:
            processed_df['department'] = processed_df[department_col]
        elif 'department' not in processed_df.columns:
            # 如果没有部门信息，使用工作表名称作为部门
            processed_df['department'] = sheet_name
                    
        # 添加工作表名称
        processed_df['sheet_name'] = sheet_name
        
        # 数据清理
        if 'position_name' in processed_df.columns:
            try:
                processed_df['position_name'] = processed_df['position_name'].astype(str).str.strip()
                # 移除空值和无效数据
                processed_df = processed_df[processed_df['position_name'].notna()]
                processed_df = processed_df[processed_df['position_name'] != '']
                processed_df = processed_df[processed_df['position_name'] != 'nan']
            except Exception as e:
                self.logger.error(f"处理职位名称列时发生错误: {e}")
                processed_df = processed_df[processed_df['position_name'].notna()]
            
        if 'position_code' in processed_df.columns:
            try:
                processed_df['position_code'] = processed_df['position_code'].astype(str).str.strip()
            except Exception as e:
                self.logger.error(f"处理职位代码列时发生错误: {e}")
            
        if 'department' in processed_df.columns:
            try:
                processed_df['department'] = processed_df['department'].astype(str).str.strip()
            except Exception as e:
                self.logger.error(f"处理部门列时发生错误: {e}")
        else:
            # 如果没有部门信息，使用工作表名称作为部门
            processed_df['department'] = sheet_name
            
        # 重置索引
        processed_df = processed_df.reset_index(drop=True)
        
        self.logger.info(f"工作表 '{sheet_name}' 预处理完成，保留 {len(processed_df)} 条有效记录")
        return processed_df
        
    def extract_position_info(self, position_sheets: Dict[str, pd.DataFrame]) -> List[Dict]:
        """
        从职位表数据中提取职位信息
        
        Args:
            position_sheets: 职位表工作表数据
            
        Returns:
            List[Dict]: 职位信息列表
        """
        positions = []
        
        for sheet_name, df in position_sheets.items():
            for _, row in df.iterrows():
                position_info = {
                    'sheet_name': sheet_name,
                    'position_name': row.get('position_name', ''),
                    'position_code': row.get('position_code', ''),
                    'department': row.get('department', sheet_name),
                    'row_index': row.name
                }
                
                # 确保职位名称不为空
                if position_info['position_name']:
                    positions.append(position_info)
                    
        self.logger.info(f"提取职位信息完成，共 {len(positions)} 个职位")
        return positions
        
    def validate_interview_data_structure(self, df: pd.DataFrame, sheet_name: str = "面试名单") -> bool:
        """
        验证面试人员名单表数据结构
        
        Args:
            df: 数据框
            sheet_name: 工作表名称
            
        Returns:
            bool: 数据结构是否有效
            
        Raises:
            ExcelProcessingError: 数据结构无效时抛出异常
        """
        if df.empty:
            raise ExcelProcessingError(f"工作表 '{sheet_name}' 没有数据")
            
        # 检查必需的列
        required_columns = {
            'name': ['姓名', '名字', '面试人员', '候选人'],
            'position': ['岗位', '职位', '岗位名称', '职位名称', '报考岗位'],
            'score': ['分数', '成绩', '笔试成绩', '面试成绩', '总分']
        }
        
        found_columns = {}
        
        for col_type, keywords in required_columns.items():
            found_column = None
            for col in df.columns:
                col_str = str(col).strip()
                if any(keyword in col_str for keyword in keywords):
                    found_column = col
                    break
            
            if found_column is None:
                available_columns = list(df.columns)
                raise ExcelProcessingError(
                    f"工作表 '{sheet_name}' 缺少{col_type}相关列。"
                    f"需要包含包含以下关键词的列: {keywords}，"
                    f"当前列: {available_columns}"
                )
            
            found_columns[col_type] = found_column
            
        # 检查数据是否有效
        name_col = found_columns['name']
        position_col = found_columns['position']
        score_col = found_columns['score']
        
        valid_names = df[name_col].dropna()
        valid_positions = df[position_col].dropna()
        valid_scores = pd.to_numeric(df[score_col], errors='coerce').dropna()
        
        if len(valid_names) == 0:
            raise ExcelProcessingError(f"工作表 '{sheet_name}' 的姓名列没有有效数据")
        if len(valid_positions) == 0:
            raise ExcelProcessingError(f"工作表 '{sheet_name}' 的岗位列没有有效数据")
        if len(valid_scores) == 0:
            raise ExcelProcessingError(f"工作表 '{sheet_name}' 的分数列没有有效数据")
            
        self.logger.info(f"工作表 '{sheet_name}' 数据结构验证通过，找到 {len(valid_names)} 个面试人员")
        return True
        
    def read_interview_file(self, file_path: str) -> pd.DataFrame:
        """
        读取面试人员名单表文件
        
        Args:
            file_path: 面试人员名单表文件路径
            
        Returns:
            pd.DataFrame: 面试人员数据
            
        Raises:
            ExcelProcessingError: 读取失败时抛出异常
        """
        self.validate_file_path(file_path)
        self.check_excel_format(file_path)
        
        try:
            # 读取第一个工作表（通常面试名单只有一个工作表）
            df = self.read_excel_sheet(file_path)
            
            # 验证数据结构
            self.validate_interview_data_structure(df)
            
            # 数据预处理
            processed_df = self._preprocess_interview_data(df)
            
            self.logger.info(f"面试人员名单读取完成，包含 {len(processed_df)} 个面试人员")
            return processed_df
            
        except Exception as e:
            if isinstance(e, ExcelProcessingError):
                raise
            else:
                raise ExcelProcessingError(f"读取面试人员名单文件时发生错误: {str(e)}")
                
    def _preprocess_interview_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        预处理面试人员数据
        
        Args:
            df: 原始数据框
            
        Returns:
            pd.DataFrame: 预处理后的数据框
        """
        try:
            # 创建数据副本
            processed_df = df.copy()
            
            # 确保没有重复的列名
            if processed_df.columns.duplicated().any():
                self.logger.warning("发现重复的列名，正在处理...")
                # 为重复的列名添加后缀
                cols = processed_df.columns.tolist()
                seen = {}
                for i, col in enumerate(cols):
                    if col in seen:
                        seen[col] += 1
                        cols[i] = f"{col}_{seen[col]}"
                    else:
                        seen[col] = 0
                processed_df.columns = cols
            
            # 标准化列名映射
            column_mapping = {}
            name_col = None
            position_col = None
            score_col = None
            
            for col in processed_df.columns:
                col_str = str(col).strip()
                
                # 姓名列
                if not name_col and any(keyword in col_str for keyword in ['姓名', '名字', '面试人员', '候选人']):
                    name_col = col
                    column_mapping[col] = 'name'
                # 岗位列
                elif not position_col and any(keyword in col_str for keyword in ['岗位', '职位', '报考岗位', '招考职位']):
                    position_col = col
                    column_mapping[col] = 'position_name'
                # 分数列
                elif not score_col and any(keyword in col_str for keyword in ['分数', '成绩', '总分', '面试分数']):
                    score_col = col
                    column_mapping[col] = 'score'
            
            # 检查是否找到了必需的列
            if not name_col:
                raise ExcelProcessingError("未找到姓名列")
            if not position_col:
                raise ExcelProcessingError("未找到职位列")
            if not score_col:
                raise ExcelProcessingError("未找到分数列")
            
            # 添加标准化列名，但保留原始列名
            # 这样既有标准化的列名（供传统匹配器使用），又有原始列名（供可配置匹配器使用）
            if name_col and 'name' not in processed_df.columns:
                processed_df['name'] = processed_df[name_col]
            
            if position_col and 'position_name' not in processed_df.columns:
                processed_df['position_name'] = processed_df[position_col]
            
            if score_col and 'score' not in processed_df.columns:
                processed_df['score'] = processed_df[score_col]
            
            # 数据清理和验证
            if 'name' in processed_df.columns:
                try:
                    # 安全地处理姓名列
                    name_series = processed_df['name'].astype(str).str.strip()
                    processed_df['name'] = name_series
                    processed_df = processed_df[processed_df['name'].notna()]
                    processed_df = processed_df[processed_df['name'] != '']
                    processed_df = processed_df[processed_df['name'] != 'nan']
                except Exception as e:
                    self.logger.error(f"处理姓名列时发生错误: {e}")
                    # 如果处理失败，至少保留非空数据
                    processed_df = processed_df[processed_df['name'].notna()]
                
            if 'position_name' in processed_df.columns:
                try:
                    # 安全地处理职位名称列
                    position_series = processed_df['position_name'].astype(str).str.strip()
                    processed_df['position_name'] = position_series
                    processed_df = processed_df[processed_df['position_name'].notna()]
                    processed_df = processed_df[processed_df['position_name'] != '']
                    processed_df = processed_df[processed_df['position_name'] != 'nan']
                except Exception as e:
                    self.logger.error(f"处理职位名称列时发生错误: {e}")
                    # 如果处理失败，至少保留非空数据
                    processed_df = processed_df[processed_df['position_name'].notna()]
                
            if 'score' in processed_df.columns:
                # 清理分数数据
                processed_df = self._clean_score_data(processed_df)
                
            # 添加资格标记（假设60分及格）
            if 'score' in processed_df.columns:
                processed_df['is_qualified'] = processed_df['score'] >= 60
                
            # 重置索引
            processed_df = processed_df.reset_index(drop=True)
            
            self.logger.info(f"面试数据预处理完成，保留 {len(processed_df)} 条有效记录")
            return processed_df
            
        except Exception as e:
            self.logger.error(f"面试数据预处理失败: {e}")
            raise ExcelProcessingError(f"面试数据预处理失败: {str(e)}")
        
    def _clean_score_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        清理分数数据，处理无效分数
        
        Args:
            df: 包含分数列的数据框
            
        Returns:
            pd.DataFrame: 清理后的数据框
        """
        if 'score' not in df.columns:
            return df
            
        # 转换分数为数值类型
        df['score'] = pd.to_numeric(df['score'], errors='coerce')
        
        # 记录无效分数的数量
        invalid_scores = df['score'].isna().sum()
        if invalid_scores > 0:
            self.logger.warning(f"发现 {invalid_scores} 个无效分数，将被移除")
            
        # 移除无效分数的行
        df = df.dropna(subset=['score'])
        
        # 验证分数范围（只移除负分数，不限制上限）
        invalid_range = df[df['score'] < 0]
        if len(invalid_range) > 0:
            self.logger.warning(f"发现 {len(invalid_range)} 个负分数，将被移除")
            df = df[df['score'] >= 0]
            
        return df
        
    def extract_interview_info(self, interview_df: pd.DataFrame) -> List[Dict]:
        """
        从面试人员数据中提取面试信息
        
        Args:
            interview_df: 面试人员数据框
            
        Returns:
            List[Dict]: 面试人员信息列表
        """
        interviews = []
        
        for _, row in interview_df.iterrows():
            interview_info = {
                'name': row.get('name', ''),
                'position_name': row.get('position_name', ''),
                'score': row.get('score', 0.0),
                'is_qualified': row.get('is_qualified', False),
                'row_index': row.name
            }
            
            # 确保必要字段不为空
            if interview_info['name'] and interview_info['position_name']:
                interviews.append(interview_info)
                
        self.logger.info(f"提取面试信息完成，共 {len(interviews)} 个面试人员")
        return interviews
        
    def get_position_candidates(self, interview_df: pd.DataFrame, position_name: str) -> List[Dict]:
        """
        获取指定岗位的所有面试人员
        
        Args:
            interview_df: 面试人员数据框
            position_name: 岗位名称
            
        Returns:
            List[Dict]: 该岗位的面试人员列表
        """
        if 'position_name' not in interview_df.columns:
            return []
            
        # 筛选指定岗位的面试人员
        position_candidates = interview_df[interview_df['position_name'] == position_name]
        
        candidates = []
        for _, row in position_candidates.iterrows():
            candidate_info = {
                'name': row.get('name', ''),
                'score': row.get('score', 0.0),
                'is_qualified': row.get('is_qualified', False)
            }
            candidates.append(candidate_info)
            
        return candidates
        
    def get_min_score_by_position(self, interview_df: pd.DataFrame) -> Dict[str, float]:
        """
        计算每个岗位的最低分数
        
        Args:
            interview_df: 面试人员数据框
            
        Returns:
            Dict[str, float]: 岗位名称到最低分数的映射
        """
        if 'position_name' not in interview_df.columns or 'score' not in interview_df.columns:
            return {}
            
        # 按岗位分组计算最低分数
        min_scores = interview_df.groupby('position_name')['score'].min().to_dict()
        
        self.logger.info(f"计算最低分数完成，涉及 {len(min_scores)} 个岗位")
        return min_scores