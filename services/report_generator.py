"""
报告生成器
负责生成Excel格式的岗位分数汇总报告
"""
import pandas as pd
from datetime import datetime
from typing import List, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from models.data_models import PositionScoreResult
from services.file_manager import FileManager


class ReportGenerator:
    """Excel报告生成器"""
    
    def __init__(self):
        self.report_title = "岗位最低进面分数汇总报告"
        self.generated_time = None
        self.file_manager = FileManager()
    
    def generate_report(self, results: List[PositionScoreResult], output_path: str, 
                       conflict_strategy: str = "auto_rename") -> Tuple[bool, str]:
        """
        生成Excel格式的汇总报告
        
        Args:
            results: 岗位分数结果列表
            output_path: 输出文件路径
            conflict_strategy: 文件冲突处理策略 ("auto_rename", "overwrite", "backup")
            
        Returns:
            Tuple[bool, str]: (生成是否成功, 实际保存的文件路径)
        """
        try:
            # 处理保存路径
            processed_path = self.file_manager.get_save_path(output_path)
            
            # 验证路径有效性
            is_valid, error_msg = self.file_manager.validate_save_path(processed_path)
            if not is_valid:
                raise Exception(f"保存路径无效: {error_msg}")
            
            # 处理文件冲突
            final_path = self.file_manager.handle_file_conflict(processed_path, conflict_strategy)
            
            # 确保目录存在
            if not self.file_manager.ensure_directory_exists(final_path):
                raise Exception("无法创建输出目录")
            
            self.generated_time = datetime.now()
            
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "岗位分数汇总"
            
            # 添加报告头部
            self._add_report_header(ws)
            
            # 添加数据表格
            self._add_data_table(ws, results)
            
            # 应用样式
            self._apply_styles(ws, len(results))
            
            # 保存文件
            wb.save(final_path)
            return True, final_path
            
        except Exception as e:
            raise Exception(f"生成报告失败: {str(e)}")
    
    def _add_report_header(self, worksheet):
        """添加报告头部信息"""
        # 报告标题
        worksheet['A1'] = self.report_title
        worksheet.merge_cells('A1:I1')
        
        # 生成时间
        worksheet['A2'] = f"生成时间: {self.generated_time.strftime('%Y-%m-%d %H:%M:%S')}"
        worksheet.merge_cells('A2:I2')
        
        # 空行
        worksheet['A3'] = ""
        
        # 表头 - 包含岗位表中的重要列
        headers = ['岗位代码', '岗位名称', '用人司局', '部门名称', '招考人数', '面试人数', '面试分数', '状态', '备注']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col)
            cell.value = header
    
    def _add_data_table(self, worksheet, results: List[PositionScoreResult]):
        """添加数据表格"""
        start_row = 5
        
        for row_idx, result in enumerate(results, start_row):
            worksheet.cell(row=row_idx, column=1, value=result.position_code)
            worksheet.cell(row=row_idx, column=2, value=result.position_name)
            worksheet.cell(row=row_idx, column=3, value=result.department)  # 用人司局
            worksheet.cell(row=row_idx, column=4, value=result.department_name)  # 部门名称
            worksheet.cell(row=row_idx, column=5, value=result.recruit_count)  # 招考人数
            worksheet.cell(row=row_idx, column=6, value=result.candidate_count)  # 面试人数
            # 面试分数 - 显示所有分数而不是最低分
            scores_text = ', '.join(map(str, result.all_scores)) if result.all_scores else "无数据"
            worksheet.cell(row=row_idx, column=7, value=scores_text)
            worksheet.cell(row=row_idx, column=8, value=result.status)
            worksheet.cell(row=row_idx, column=9, value=result.notes)
    
    def _apply_styles(self, worksheet, data_rows: int):
        """应用样式格式"""
        # 标题样式
        title_cell = worksheet['A1']
        title_cell.font = Font(name='微软雅黑', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 时间样式
        time_cell = worksheet['A2']
        time_cell.font = Font(name='微软雅黑', size=10)
        time_cell.alignment = Alignment(horizontal='center')
        
        # 表头样式
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        
        for col in range(1, 10):  # A到I列
            cell = worksheet.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行样式
        data_font = Font(name='微软雅黑', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用到所有数据单元格
        for row in range(4, 5 + data_rows):  # 包括表头和数据行
            for col in range(1, 10):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                if row > 4:  # 数据行
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 调整列宽
        column_widths = [12, 25, 15, 15, 12, 12, 30, 15, 20]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    def create_summary_statistics(self, results: List[PositionScoreResult]) -> dict:
        """创建汇总统计信息"""
        total_positions = len(results)
        normal_positions = len([r for r in results if r.status == "正常"])
        no_interview_positions = len([r for r in results if r.status == "无面试人员"])
        error_positions = len([r for r in results if r.status == "数据异常"])
        
        return {
            "total_positions": total_positions,
            "normal_positions": normal_positions,
            "no_interview_positions": no_interview_positions,
            "error_positions": error_positions,
            "success_rate": f"{(normal_positions / total_positions * 100):.1f}%" if total_positions > 0 else "0%"
        }
    
    def get_recommended_save_path(self, base_directory: Optional[str] = None) -> str:
        """
        获取推荐的保存路径
        
        Args:
            base_directory: 基础目录，如果不指定则使用当前目录
            
        Returns:
            str: 推荐的保存路径
        """
        if base_directory:
            return self.file_manager.get_save_path(base_directory)
        else:
            return self.file_manager.get_save_path()
    
    def validate_output_path(self, output_path: str) -> Tuple[bool, str]:
        """
        验证输出路径
        
        Args:
            output_path: 要验证的输出路径
            
        Returns:
            Tuple[bool, str]: (是否有效, 错误信息)
        """
        processed_path = self.file_manager.get_save_path(output_path)
        return self.file_manager.validate_save_path(processed_path)